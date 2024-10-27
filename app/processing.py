import csv
import os

from openpyxl import load_workbook


def load_mappings(mapping_file):
    mappings = {}
    with open(mapping_file, mode='r') as file:
        csv_reader = csv.reader(file)
        for row in csv_reader:
            input_cell, output_cell = row[0], row[1]
            mappings[input_cell] = output_cell
    return mappings


def process_file(input_file, template_file, mappings):
    # Load input Excel file
    input_wb = load_workbook(input_file)
    input_ws = input_wb['Card']  # Select sheet by name

    # Load template Excel file
    template_wb = load_workbook(template_file)
    template_ws = template_wb['Main']  # Select sheet by name

    # Apply mappings: copy from input to template
    for input_cell, output_cell in mappings.items():
        input_value = input_ws[input_cell].value
        template_ws[output_cell].value = input_value

    return template_wb


def save_output(output_folder, output_file_name, workbook):
    output_path = os.path.join(output_folder, output_file_name)
    workbook.save(output_path)


def process_all_files(input_folder, template_file, mappings, output_folder, append_text=""):
    input_files = [f for f in os.listdir(input_folder) if f.endswith(('.xlsx', '.xls', 'xlsm'))]

    for input_file in input_files:
        input_path = os.path.join(input_folder, input_file)

        # Process the input file with the template
        processed_workbook = process_file(input_path, template_file, mappings)

        # Generate output file name (original name + append text)
        output_file_name = f"{os.path.splitext(input_file)[0]}{append_text}.xlsx"

        # Save the processed file to the output folder
        save_output(output_folder, output_file_name, processed_workbook)

        print(f"Processed: {input_file} -> {output_file_name}")


def run_processing(input_folder, template_file, mapping_file, output_folder, append_text="_processed"):
    mappings = load_mappings(mapping_file)
    process_all_files(input_folder, template_file, mappings, output_folder, append_text)
